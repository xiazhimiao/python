import time
from selenium import webdriver
from selenium.common import ElementClickInterceptedException, TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from pathlib import Path


def setup_driver():
    """初始化浏览器驱动"""
    service = Service(executable_path="./chromedriver.exe")  # 替换为实际路径
    driver = webdriver.Chrome(service=service)
    driver.maximize_window()
    return driver


def search_major(driver):
    """搜索专业"""
    driver.get('https://yz.chsi.com.cn/zsml/')
    print("🔍 正在搜索专业：计算机科学与技术")
    input_box = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='请输入专业名称']"))
    )
    input_box.send_keys("计算机科学与技术")
    search_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(@class, 'ivu-btn-primary') and .//span[text()='查询']]"))
    )
    driver.execute_script("arguments[0].click();", search_btn)  # JS点击


def select_major(driver):
    """选择专业"""
    try:
        results = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//div[@class='zy-item']"))
        )
        print("\n📑 找到以下专业：")
        for i, result in enumerate(results, 1):
            code = result.find_element(By.XPATH, ".//div[@class='zy-name']").text.split()[0]
            print(f"{i}. {code} ")

        while True:
            try:
                choice = int(input("请选择专业（数字）："))
                if 1 <= choice <= len(results):
                    break
                print(f"请输入 1 到 {len(results)} 之间的数字")
            except ValueError:
                print("请输入有效的数字")

        print(f"👉 选择第 {choice} 个专业")
        selected_major = results[choice - 1].find_element(By.XPATH, ".//div[@class='zy-name']").text.split()[0]

        # 滚动到元素并JS点击
        driver.execute_script("arguments[0].scrollIntoView();", results[choice - 1])
        select_btn = results[choice - 1].find_element(By.XPATH, ".//a[contains(@class, 'zy-btn')]")
        driver.execute_script("arguments[0].click();", select_btn)
        return selected_major
    except Exception as e:
        print(f"❗ 选择专业失败: {str(e)}")
        driver.quit()
        exit()


def switch_window_and_login(driver):
    """切换窗口并等待手动登录"""
    try:
        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])
        input("⏳ 请手动登录后按回车继续...")
        print("请在 10 秒内切换回浏览器窗口。")
        for i in range(10, 0, -1):
            print(f"剩余时间：{i} 秒")
            time.sleep(1)
    except Exception as e:
        print(f"❗ 登录窗口处理失败: {str(e)}")
        driver.quit()
        exit()


def expand_schools(driver):
    """展开当前页所有院校"""
    print("\n🔄 展开当前页院校：")
    schools = WebDriverWait(driver, 20).until(
        EC.presence_of_all_elements_located(
            (By.XPATH, "//div[@class='zy-item' and not(contains(@style, 'display: none'))]"))
    )
    for idx, school in enumerate(schools, 1):
        try:
            icon = school.find_element(By.XPATH, './/a[@class="show-more"]/i').get_attribute("class")
            if "arrows-down" in icon:
                expand_btn = WebDriverWait(school, 5).until(
                    EC.element_to_be_clickable((By.XPATH, './/a[@class="show-more"]'))
                )
                driver.execute_script("arguments[0].click();", expand_btn)
                print(
                    f"✅ 展开 {idx}/{len(schools)}: {school.find_element(By.XPATH, './/div[@class=\"yx-name\"]').text}")
                time.sleep(1)  # 点击展开后延迟 1 秒
            else:
                print(f"ℹ️ 跳过 {idx}（已展开）: {school.find_element(By.XPATH, './/div[@class=\"yx-name\"]').text}")
        except Exception as e:
            print(f"⚠️ 展开失败 {idx}: {str(e)[:30]}")
    return schools


def extract_subjects(schools, driver):
    """修复科目按钮定位 + 增强等待逻辑"""
    global school_name
    data = []
    for school in schools:
        try:
            school_name = WebDriverWait(school, 5).until(
                EC.visibility_of_element_located((By.XPATH, './/div[@class="yx-name"]'))
            ).text.strip()

            while True:
                try:
                    table = WebDriverWait(school, 10).until(
                        EC.presence_of_element_located((By.XPATH, './/div[@class="ivu-table-body"]'))
                    )
                    driver.execute_script("arguments[0].scrollIntoView();", table)
                    rows = table.find_elements(By.XPATH, "./table/tbody/tr")

                    for row in rows:
                        try:
                            department = row.find_element(By.XPATH, ".//td[1]").text.strip()
                            major = row.find_element(By.XPATH, ".//td[3]").text.strip()
                            direction = row.find_element(By.XPATH, ".//td[5]").text.strip()

                            kskm_list = WebDriverWait(row, 5).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, '.kskm-modal.kskm-detail-list'))
                            )

                            # 提取科目
                            # kskm_list = kskm.find_element(By.CSS_SELECTOR, '.kskm-detail-list')

                            subjects = []
                            for item in kskm_list.find_elements(By.CSS_SELECTOR, '.kskm-detail .item'):
                                text = item.get_attribute('innerHTML')
                                if '<span' in text:
                                    text = text.split('<span')[0].strip()
                                subjects.append(text)
                            exam_subjects = subjects[:4] + ['未公布'] * (4 - len(subjects))
                            row_data = [school_name, department, major, direction] + exam_subjects
                            data.append(row_data)

                        except Exception as e:
                            print(f"⚠️ 行提取失败（院校: {school_name}）: {str(e)}")
                            continue

                    # 分页处理
                    next_page_xpath = ".//div[@class='margin-top-12 clearfix']/ul[@class='ivu-page mini']/li[@title='下一页' and contains(@class, 'ivu-page-next')]"
                    if not click_next_page(driver,school, next_page_xpath, 1, True):
                        print("已经是最后一页，退出循环")
                        break

                except TimeoutException:
                    print(f"⚠️ 院校 {school_name} 表格加载超时，跳过")
                    break

        except Exception as e:
            print(f"⚠️ 跳过院校 {school_name}：{str(e)}")
            continue
    return data


def save_to_excel(data, major_code):
    """将数据保存到 Excel 文件（追加模式）"""
    file_path = f"{major_code}_考试科目.xlsx"
    file_exists = Path(file_path).exists()

    if file_exists:
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["院校", "院系所", "专业", "研究方向", "政治", "外语", "业务课一", "业务课二"])

    for row in data:
        ws.append(row)

    wb.save(file_path)
    print(f"数据已成功保存到 {file_path} 文件中。")


def click_next_page(driver, context, next_page_btn_selector, pause_time, use_xpath=False):
    try:
        locator_strategy = By.XPATH if use_xpath else By.CSS_SELECTOR
        next_page_btn = WebDriverWait(context, 5).until(
            EC.presence_of_element_located((locator_strategy, next_page_btn_selector))
        )

        if 'ivu-page-disabled' in next_page_btn.get_attribute('class'):
            print("下一页按钮已禁用")
            return False

        # 滚动到按钮位置，使用 driver 执行 JavaScript
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
                              next_page_btn)
        time.sleep(1)  # 等待滚动完成

        # 尝试点击按钮
        try:
            next_page_btn.click()
            print("已成功点击下一页按钮")
        except ElementClickInterceptedException:
            # 如果被其他元素遮挡，使用 driver 执行 JavaScript 点击
            driver.execute_script("arguments[0].click();", next_page_btn)
            print("使用 JS 点击成功")

        # 倒计时等待，使用传入的暂停时间
        for i in range(pause_time, 0, -1):
            print(f"等待 {i} 秒...")
            time.sleep(1)

        return True

    except TimeoutException:
        print("未找到下一页按钮")
        return False


def main():
    start_time = time.time()
    driver = setup_driver()
    try:
        search_major(driver)
        selected_major = select_major(driver)
        switch_window_and_login(driver)

        page_num = 1  # 初始化页码
        while True:
            print(f"正在处理第 {page_num} 页")
            schools = expand_schools(driver)
            page_data = extract_subjects(schools, driver)

            save_to_excel(page_data, selected_major)

            # 定位全局分页按钮
            next_page_xpath = "//div[@class='margin-top-20 clearfix']/ul[@class='ivu-page']/li[@title='下一页' and contains(@class, 'ivu-page-next')]"
            if not click_next_page(driver,driver, next_page_xpath, 5, True):
                print("全局下一页按钮不可用，终止")
                break

            page_num += 1

    except KeyboardInterrupt:
        print("\n程序被用户中断")
    except Exception as e:
        print(f"\n❌ 致命错误：{str(e)}")
        driver.save_screenshot("error.png")
    finally:
        driver.quit()
    end_time = time.time()
    print(f"程序运行时间: {end_time - start_time:.2f} 秒")
    print(
        "提示：使用Selenium时，尽量保持浏览器窗口为激活状态，避免切屏影响元素加载和交互。若仍有问题，可适当增加关键步骤的等待时间。")


if __name__ == "__main__":
    main()
